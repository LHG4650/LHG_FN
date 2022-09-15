import cv2
import os
import time
import openpyxl

def fh2_img_const_zero_bler(img,cosnt_alpha=10):
    #_, img = cv2.threshold(img,70,255, cv2.THRESH_TOZERO)

    lab = cv2.cvtColor(img, cv2.COLOR_BGR2LAB)
    l, a, b = cv2.split(lab)
    clahe = cv2.createCLAHE(clipLimit=cosnt_alpha,tileGridSize=(8, 8))
    l = clahe.apply(l)
    lab = cv2.merge((l, a, b))
    cont_dst = cv2.cvtColor(lab, cv2.COLOR_LAB2BGR)

    cont2blur10 = cv2.fastNlMeansDenoisingColored(cont_dst,None,5,5,7,21)

    return cont2blur10

def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print ('Error: Creating directory. ' +  directory)



def set_log(fh_no):      #로그파일생성, output : 로그파일 생성 결과(1,0), 로그파일 경로
    time_stamp = time.strftime('%y.%m.%d_%H.%M',time.localtime()) #현재시간 yy>mm.dd
    file_dir = "C:/HG_001_Log/fh_"+str(fh_no)
    save_file_path = file_dir + "/" + time_stamp + ".xlsx"
    createFolder(file_dir)
    wb=openpyxl.Workbook()
    ws = wb.active
    ws["A1"].value = "time"
    ws["B1"].value = 'value_dist_fh1'
    ws["C1"].value = 'value_dist_fh2'
    wb.save(save_file_path)
    return save_file_path

def write_log(log_file,value,value2):
    time_stamp = time.strftime('%y.%m.%d_%H.%M.%S',time.localtime()) #현재시간 yy>mm.dd
    wb = openpyxl.load_workbook(log_file)
    ws = wb.active
    len_log = len(ws['a'])
    ws.cell(row = int(len_log+1), column=1).value = time_stamp
    ws.cell(row = int(len_log+1), column=2).value = value
    ws.cell(row = int(len_log+1), column=3).value = value2
    wb.save(log_file)